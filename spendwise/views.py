from django.shortcuts import render
from spendwise.models import Transaction
from datetime import datetime
import openpyxl

def index(request):
    if request.method == "GET":
        return render(request, 'spendwise/index.html')
    else:
        if not request.user.is_authenticated:
            return render(request, 'spendwise/index.html')

        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)

        # get first sheet
        worksheet = wb[wb.sheetnames[0]]

        # TODO make sure the worksheet is in correct format
        # delete unnecessary rows
        worksheet.delete_rows(0, 3)

        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            date = datetime.strptime(row[0].value,"%Y.%m.%d %H:%M").replace(tzinfo=None)
            description = row[2].value
            amount = int(row[4].value) - int(row[3].value)
            transaction = Transaction.objects.create(
                user=request.user,
                datetime=date,
                description=description,
                amount=amount)
            transaction.find_place()
            transaction.save()
        return render(request, 'spendwise/index.html')

def txs(request):
    if request.user.is_authenticated:
        user_transactions = Transaction.objects.filter(user=request.user).order_by('-datetime')
        print(user_transactions.count())
        context = {'txs': user_transactions}
        return render(request, 'spendwise/txs.html', context)
    else:
        return index(request)
